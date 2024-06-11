import logging
import os
import time
import traceback
from typing import Optional, List

import openpyxl
from django.http import HttpRequest, HttpResponse, FileResponse
from django.utils.encoding import escape_uri_path
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from util import utils
from util.restful import RestResponse
from web_app.dao import user_dao
from web_app.decorators.admin_decorator import log_func
from web_app.model.link import AccountLink2
from web_app.model.users import User, USER_ROLE_BUSINESS, USER_ROLE_ADMIN, USER_ROLE_UPLOADER
from web_app.service import link_service
from util.exception import BusinessException
from web_app.settings import BASE_DIR


@log_func
def page_list(request: HttpRequest):
    start_row, end_row = utils.page_query(request)
    body = utils.request_body(request)
    try:
        res, count = link_service.query_list_page_new(AccountLink2.objects, body, start_row, end_row)
    except:
        logging.info("linkPageList# err = %s", traceback.format_exc())
        res = []
        count = 0
    return RestResponse.success_list(count=count, data=res)


@log_func
def add_data(request: HttpRequest):
    body = utils.request_body(request)
    user = user_dao.get_user(request)
    if not user:
        return RestResponse.failure("操作用户不存在")
    if user.role == USER_ROLE_UPLOADER:
        return RestResponse.failure("非业务员不可操作")
    try:
        link_service.insert_new(AccountLink2.objects, body, user)
    except BusinessException as e:
        return RestResponse.failure(e.errmsg)
    return RestResponse.success("添加成功")


@log_func
def update_data(request: HttpRequest):
    body = utils.request_body(request)
    user = user_dao.get_user(request)
    if not user:
        return RestResponse.failure("操作用户不存在")
    if user.role == USER_ROLE_UPLOADER:
        return RestResponse.failure("非业务员不可操作")
    try:
        link_service.update_new(AccountLink2.objects, body, user)
    except BusinessException as e:
        return RestResponse.failure(e.errmsg)
    return RestResponse.success("修改成功")


@log_func
def delete_data(request: HttpRequest):
    body = utils.request_body(request)
    user = user_dao.get_user(request)
    if not user:
        return RestResponse.failure("操作用户不存在")
    if user.role == USER_ROLE_UPLOADER:
        return RestResponse.failure("非业务员不可操作")
    link_service.delete_link(AccountLink2.objects, body)
    return RestResponse.success("成功")


TEMP_DIR = os.path.join(BASE_DIR, "data", 'temp')


def export(request: HttpRequest):
    body = utils.request_body(request)
    query_list = AccountLink2.objects.values(
        'id', "link", "remark", "op_user__username",
        "create_time", "update_time"
    )

    data_size = len(query_list)
    if data_size <= 0:
        return HttpResponse(status=404, content="下载失败，没有数据")

    limit = 1000
    count = int(data_size / limit) + 1
    logging.info("link_export id # 数据大小 = %s, count = %s", data_size, count)
    file_path = None
    for index in range(0, count):
        offset = index * limit
        _next = offset + limit
        data_list = query_list[offset:_next]
        logging.info("offset = %s, next = %s size = %s", offset, _next, len(data_list))
        if len(data_list) == 0:
            break

        excel_list = data_list
        logging.info("link_export# index = %s 转换ExcelBean成功 = %s", index, len(excel_list))
        if len(excel_list) <= 0:
            continue
        if not file_path:
            temp_path = os.path.join(TEMP_DIR, "excel")
            if not os.path.exists(temp_path):
                os.mkdir(temp_path)
            logging.info("link_export#开始生成excel文件，目录 = %s", temp_path)
            file_path = create_excel(excel_list, temp_path)
            if not file_path:
                return HttpResponse(status=404, content="下载失败，创建excel失败")
        else:
            logging.info("wa_account_id2_export# index = %s追加生成excel文件，文件 = %s", index, file_path)
            append_excel(excel_list, file_path)

    logging.info("link_export#数据填充完毕，最终文件 = %s", file_path)
    file = open(file_path, 'rb')
    file_name = os.path.basename(file_path)
    response = FileResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename="' + escape_uri_path(file_name) + '"'
    return response


HEADER_OPTIONS = [
    {'title': '链接', 'col_width': 80},
    {'title': '备注', 'col_width': 30, },
    {'title': '上传人', 'col_width': 10, },
    {'title': '上传时间', 'col_width': 10, },
]


def create_excel(data_list: List[dict], out_dir: str) -> Optional[str]:
    try:
        book = Workbook()
        now = time.strftime("%Y-%m-%d-%H_%M_%S", time.localtime(time.time()))
        filename = os.path.join(out_dir, "链接表_" + now + ".xlsx")
        sheet: Optional[Worksheet] = book.active
        for index, head in enumerate(HEADER_OPTIONS):
            t = head.get('title')
            w = head.get('col_width')
            col_name = get_column_letter(index + 1)
            col = sheet.column_dimensions[col_name]
            col.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            col.auto_size = True
            col.width = w
            sheet.cell(1, index + 1).value = t

        for index, d in enumerate(data_list):
            try:
                row = [d.get('id'), d.get('link'), d.get('remark'), d.get('op_user__username'), d.get('create_time')]
                sheet.append(row)
            except:
                pass
        book.save(filename)
        return filename
    except:
        logging.info("create_excel fail = %s", traceback.format_exc())
    return None


def append_excel(in_data_list: List[dict], path: str):
    try:
        book = openpyxl.load_workbook(filename=path)
        sheet: Optional[Worksheet] = book.active
        for d in in_data_list:
            row = [d.get('id'), d.get('link'), d.get('remark'), d.get('op_user__username'), d.get('create_time')]
            sheet.append(row)
        book.save(filename=path)
    except BaseException:
        logging.info("append excel fail = %s", traceback.format_exc())
    return path
