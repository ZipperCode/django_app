import logging
from dataclasses import dataclass
import os.path
import time
from typing import List, Optional

from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.writer.excel import ExcelWriter


@dataclass
class ExcelBean:
    id: str = None
    qr_content: str = None
    qr_code_abs_path: str = None
    country: Optional[str] = None
    age: int = 0
    work: Optional[str] = None
    mark: Optional[str] = None
    money: float = 0.0
    op_user: str = "None"
    upload_time: str = "2023-1-1"


def create_excel(in_data_list: List[ExcelBean], out_dir):
    try:
        book = Workbook()
        sheet: Optional[Worksheet] = book.active
        sheet.append(['ID', '城市', '年龄', '工作', '收入', '备注', '上传人', '上传时间'])
        r = sheet.row_dimensions[1]
        r.font = Font(bold=True)
        now = time.strftime("%Y-%m-%d-%H_%M_%S", time.localtime(time.time()))
        filename = os.path.join(out_dir, "ID表_" + now + ".xlsx")
        for d in in_data_list:
            row = [d.id, d.country, d.age, d.work, d.money, d.mark, d.op_user, d.upload_time]
            sheet.append(row)
        book.save(filename)
    except BaseException:
        logging.info("create_excel fail")
        return None
    return filename


excel2_header_options = [
    {'title': '二维码', 'col_width': 20, },
    {'title': '二维码内容', 'col_width': 30, },
    {'title': '城市', 'col_width': 10, },
    {'title': '年龄', 'col_width': 5, },
    {'title': '工作', 'col_width': 10, },
    {'title': '收入', 'col_width': 10, },
    {'title': '备注', 'col_width': 30, },
    {'title': '上传人', 'col_width': 10, },
    {'title': '上传时间', 'col_width': 10, },
]


def create_excel2(in_data_list: List[ExcelBean], out_dir):
    try:
        book = Workbook()
        now = time.strftime("%Y-%m-%d-%H_%M_%S", time.localtime(time.time()))
        filename = os.path.join(out_dir, "二维码表_" + now + ".xlsx")

        sheet: Optional[Worksheet] = book.active
        header_row = sheet.row_dimensions[0]
        for index, head in enumerate(excel2_header_options):
            t = head.get("title")
            w = head.get("col_width")
            col_name = get_column_letter(index + 1)
            col = sheet.column_dimensions[col_name]
            col.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            col.auto_size = True
            col.width = w
            sheet.cell(1, index + 1).value = t

        for index, d in enumerate(in_data_list):
            img = Image(d.qr_code_abs_path)  # 缩放图片
            img.width, img.height = (100, 100)
            _from = AnchorMarker(0, 0, index + 1, 0)
            to = AnchorMarker(1, 0, index + 2, 0)
            img.anchor = TwoCellAnchor('twoCell', _from, to)
            sheet.add_image(img)

            rows = [d.qr_content, d.country, d.age, d.work, d.money, d.mark, d.op_user, d.upload_time]

            for r_index, value in enumerate(rows):
                c = sheet.cell(index + 2, r_index + 2)
                c.value = value
                c.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            sheet.row_dimensions[index + 2].height = 100
            sheet.row_dimensions[index + 2].width = 100
        book.save(filename)
    except BaseException as e:
        logging.info("create_excel fail")
        print(' e = ', e)
        return None
    return filename


def adjust(ws):
    column_widths = []
    for i, col in enumerate(ws.iter_cols(min_col=ws.max_column, max_col=ws.max_column, min_row=ws.min_row)):
        for cell in col:
            value = cell.value
            if value is not None:
                if isinstance(value, str) is False:
                    value = str(value)
                try:
                    column_widths[i] = max(column_widths[i], len(value))
                except IndexError:
                    column_widths.append(len(value))
    for i, width in enumerate(column_widths):
        if i == 0:
            continue
        col_name = get_column_letter(ws.min_column + i)
        value = column_widths[i] + 2
        col = ws.column_dimensions[col_name]
        col.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        col.auto_size = True
        col.width = value


if __name__ == '__main__':
    data_list = [
        ExcelBean(
            qr_code_abs_path="C:\\Users\\Zipper\\Pictures\\mmexport1686967694869.jpg",
            qr_content="qr_content",
            country="中国",
            work="没工作",
            mark="我是备注"
        ),
        ExcelBean(
            qr_code_abs_path="C:\\Users\\Zipper\\Pictures\\mmexport1686967694869.jpg",
            qr_content="qr_content"
        ),
    ]
    create_excel2(data_list, "D:\\")
